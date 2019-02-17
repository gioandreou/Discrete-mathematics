'''
ANDREOU GEORGRIOS 2334 DISCRETE MATHEMATICS 2 
The following algorithm calculates the number of solutions of equations
'''

from sympy import *
import random
import time
import itertools
import openpyxl

# function for printing the whole equation
def print_first_equation(n_number):

    string_with_values=""
    for i in range(n_number,0,-1):
        if(i==n_number):
            string_with_values = str(i)+"k"+str(i) +"  "+ string_with_values
        else:
            string_with_values = str(i)+"k"+str(i) +" + "+ string_with_values
    string_with_values = string_with_values + "= " + str(n_number)
    return(string_with_values)

# function for printing the whole equation with parameters
def print_second_equation(n_number,parametersList):
    param_iter = iter(parametersList) 
    sec_string_with_values=""
    for i in range(1,n_number+1):
        if(i==1):
            sec_string_with_values = str(next(param_iter))+"l"+str(i) 
        else:
            sec_string_with_values = sec_string_with_values +" + "+ str(next(param_iter))+"l"+str(i) 
    sec_string_with_values = sec_string_with_values + " = " + str(n_number)
    return(sec_string_with_values)

# partition(): recursivly finds all the partitions of the given n_number
def partitions(n):
	# base case of recursion: zero is the sum of the empty list
	if n == 0:
		yield []
		return
		
	# modify partitions of n-1 to form partitions of n
	for p in partitions(n-1):
		yield [1] + p
		if p and (len(p) < 2 or p[1] > p[0]):
			yield [p[0] + 1] + p[1:]

# distinct_partition(): finds the partitions that contain only distinct-different digits
def distinct_partition(set_of_partitions):
    # print(set_of_partitions) #every partition 
    duplicate_partitions_remove=list()

    # find the partitions the have non-distinct digits in order to remove them 
    for element in set_of_partitions:
        list_element = list(element)
        if len(set(list_element)) != len(list_element):
            duplicate_partitions_remove.append(element)

    #delete the non-distinct partitions from the first set of partitions
    for item in duplicate_partitions_remove:
        set_of_partitions.remove(item)
       
    return set_of_partitions
    
def partition_nr_into_given_set_of_nrs(nr,S):
    nrs = sorted(S, reverse=True)
    def inner(n, i):
        if n == 0:
            yield []
        for k in range(i, len(nrs)):
            if nrs[k] <= n:
                for rest in inner(n - nrs[k], k):
                    yield [nrs[k]] + rest
    return list(inner(nr, 0))
    
    result= list(inner(nr, 0))
    uniqueList=[]
    for item in result:
        if item not in uniqueList:
            uniqueList.append(item)
    return uniqueList
    

'''
def find_solutions(unique_set,size):
    solutions=[]
    for item in unique_set:
        all_zeros = size*'0'
        #print(all_zeros)
        a_solution = list(all_zeros)
        #print(a_solution)
        for val in item:
            a_solution[val-1]=1
        solutions.append(a_solution)
    #print("Number of distinct solutions: \t{num_sol}\n".format(num_sol=len(solutions)))
    #print("Distinct solutions of equation: \t{sol}\n".format(sol=solutions))
    return solutions
'''

def printing(n_number,non_din_number,unique_set,parametersList):
    print("First Equation : \t {equation}\n".format(equation=print_first_equation(n_number)))
    print("Second Equation : \t {equation}\n".format(equation=print_second_equation(n_number,parametersList)))
    print("Number of Non Distinct solutions:\t{num_sol}\n".format(num_sol=non_din_number))
    print("Number of Distinct solutions:\t{num_sol}\n".format(num_sol=len(unique_set)))
    print("Number of solutions within Parameter set:\t{num_sol}\n".format(
        num_sol=len(partition_nr_into_given_set_of_nrs(n_number,parametersList))))

def main():
    
    n_number = int(input("Give the n number: \t"))
    #n_number=num
    parametersList = []
    answer = input("Do you want to insert the parameters (Y) or you prefer random numbers (N) ? Y/N ")
    #answer= 'n'
    if(answer == 'Y' or answer == 'y'):
        for i in range (n_number):
            temp = int(input("Give the a{a} parameter: \t".format(a=i+1)))
            parametersList.append(temp)
    else:
        low = int(input("\ngive low limit of random numbers for parameter, minimum=1\t"))
        high = int(input("\ngive high limit of random numbers for parameter, highest={a}\t".format(a=n_number)))
        for i in range(n_number):
            start=low
            end=high
            parametersList.append(random.randint(start,end))

    #print(parametersList)
    start_time = time.time()
    #create generators using the partitions method
    non_distinct_gen=partitions(n_number)
    non_dist_number = 0
    for i in non_distinct_gen: #count non distinct partitions of n number
        non_dist_number+=1

    #create a list to store all the partitions to a list called unique
    # then this list will be usesd in distinct_patrition() to keep only the 
    # distinct partitions of the n number  
    distinct_gen=partitions(n_number)
    unique=list()
    for i in distinct_gen:
        unique.append(i)

    unique_set = distinct_partition(unique)
         
    printing(n_number,non_dist_number,unique_set,parametersList)

    run_time = time.time() - start_time
    
    #writting data into excel file 
    file_xlsx = 'xlsx-descrete-an-time-n.xlsx'
    wb = openpyxl.load_workbook(filename=file_xlsx)
    ws = wb.worksheets[0]
    row = ws.max_row + 1
    ws.cell(row=row, column=1).value=n_number
    ws.cell(row=row, column=2).value=run_time
    ws.cell(row=row, column=3).value=str(start)+"-"+str(end)
    wb.save(file_xlsx)

    print("--- %s seconds ---" % run_time)

if __name__ == "__main__":
    '''
    #for multiple iterations
    for i in range(1,40,5):
        for j in range(1,i,5):
            main(i,j)
    '''
    main()